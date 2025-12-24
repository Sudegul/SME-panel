from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from typing import List, Optional
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

from ..database import get_db
from ..models.employee import Employee, EmployeeRole
from ..models.doctor_visit import DoctorVisit
from ..models.pharmacy_visit import PharmacyVisit
from ..schemas.daily_visit import (
    DoctorVisitCreate,
    DoctorVisitResponse,
    PharmacyVisitCreate,
    PharmacyVisitResponse
)
from ..utils.dependencies import get_current_user

router = APIRouter(prefix="/daily-visits", tags=["Daily Visits"])


@router.get("/doctors", response_model=List[DoctorVisitResponse])
def get_doctor_visits(
    visit_date: Optional[date] = None,
    employee_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Doktor ziyaretlerini listele
    """
    query = db.query(DoctorVisit)

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        query = query.filter(DoctorVisit.employee_id == current_user.id)
    elif employee_id:
        query = query.filter(DoctorVisit.employee_id == employee_id)

    # Tarih filtresi
    if visit_date:
        query = query.filter(DoctorVisit.visit_date == visit_date)

    visits = query.order_by(DoctorVisit.visit_date.desc()).offset(skip).limit(limit).all()
    return visits


@router.post("/doctors", response_model=DoctorVisitResponse)
def create_doctor_visit(
    visit: DoctorVisitCreate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Yeni doktor ziyareti ekle
    """
    db_visit = DoctorVisit(
        employee_id=current_user.id,
        **visit.dict()
    )

    db.add(db_visit)
    db.commit()
    db.refresh(db_visit)
    return db_visit


@router.get("/doctors/{visit_id}", response_model=DoctorVisitResponse)
def get_doctor_visit(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Belirli bir doktor ziyaretini getir
    """
    visit = db.query(DoctorVisit).filter(DoctorVisit.id == visit_id).first()

    if not visit:
        raise HTTPException(status_code=404, detail="Visit not found")

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        if visit.employee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized")

    return visit


@router.put("/doctors/{visit_id}", response_model=DoctorVisitResponse)
def update_doctor_visit(
    visit_id: int,
    visit_update: DoctorVisitCreate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Doktor ziyaretini güncelle - Sadece aynı gün 23:59'a kadar düzenlenebilir
    """
    db_visit = db.query(DoctorVisit).filter(DoctorVisit.id == visit_id).first()

    if not db_visit:
        raise HTTPException(status_code=404, detail="Visit not found")

    # Zaman kontrolü - Sadece aynı gün içinde (23:59'a kadar) düzenlenebilir
    now = datetime.now()
    visit_date_end = datetime.combine(db_visit.visit_date, datetime.max.time())  # Visit günü 23:59:59

    # EMPLOYEE ise: Sadece kendi ziyaretini ve sadece aynı gün içinde düzenleyebilir
    if current_user.role == EmployeeRole.EMPLOYEE:
        if db_visit.employee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Bu ziyaret size ait değil")

        if now > visit_date_end:
            raise HTTPException(
                status_code=403,
                detail="Ziyareti düzenlemek için son tarih geçti. Sadece ziyaret günü 23:59'a kadar düzenleyebilirsiniz."
            )

    # MANAGER/ADMIN ise: Herhangi bir ziyareti düzenleyebilir (zaman sınırı yok)
    elif current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    # Güncelle - Tüm güncellenebilir field'ları kaydet
    update_data = visit_update.dict(exclude_unset=True, exclude={'visit_date'})
    for field, value in update_data.items():
        setattr(db_visit, field, value)

    db.commit()
    db.refresh(db_visit)
    return db_visit


@router.delete("/doctors/{visit_id}")
def delete_doctor_visit(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Doktor ziyaretini sil
    """
    db_visit = db.query(DoctorVisit).filter(DoctorVisit.id == visit_id).first()

    if not db_visit:
        raise HTTPException(status_code=404, detail="Visit not found")

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        if db_visit.employee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized")

    db.delete(db_visit)
    db.commit()
    return {"message": "Visit deleted successfully"}


@router.get("/pharmacies/export")
def export_pharmacy_visits(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    employee_id: Optional[int] = None,
    pharmacy_name: Optional[str] = None,
    approval_filter: Optional[str] = None,  # 'all', 'approved', 'pending'
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Eczane ziyaretlerini Excel olarak dışa aktar (seçilen filtrelere göre)
    """
    # Yetki kontrolü - Sadece MANAGER ve ADMIN
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        raise HTTPException(
            status_code=403,
            detail="Bu işlem için yetkiniz yok"
        )

    # Query oluştur
    query = db.query(PharmacyVisit)

    # Employee filtresi
    if employee_id:
        query = query.filter(PharmacyVisit.employee_id == employee_id)

    # Eczane adı filtresi
    if pharmacy_name:
        search_term = f"%{pharmacy_name.lower().strip()}%"
        query = query.filter(PharmacyVisit.pharmacy_name.ilike(search_term))

    # Tarih aralığı filtresi
    if start_date and end_date:
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        query = query.filter(
            and_(
                PharmacyVisit.visit_date >= start_datetime,
                PharmacyVisit.visit_date <= end_datetime
            )
        )

    # Onay durumu filtresi
    if approval_filter == 'approved':
        query = query.filter(PharmacyVisit.is_approved == True)
    elif approval_filter == 'pending':
        query = query.filter(PharmacyVisit.is_approved == False)

    # Verileri çek
    visits = query.order_by(PharmacyVisit.visit_date.desc()).all()

    # Çalışan bilgisi için
    employee_name = "tum_calisanlar"
    if employee_id:
        employee = db.query(Employee).filter(Employee.id == employee_id).first()
        if employee:
            # Türkçe karakterleri düzelt ve boşlukları alt çizgi yap
            employee_name = employee.full_name.lower()
            employee_name = employee_name.replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u')
            employee_name = employee_name.replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
            employee_name = employee_name.replace(' ', '_')

    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Eczane Ziyaretleri"

    # Header stil
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Header
    headers = [
        "Çalışan", "Eczane Adı", "Adres", "Ürün Sayısı", "MF Sayısı",
        "Ziyaret Tarihi", "Başlangıç Saati", "Bitiş Saati", "Notlar", "Onay Durumu"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data
    for row_num, visit in enumerate(visits, 2):
        ws.cell(row=row_num, column=1, value=visit.employee.full_name if visit.employee else "")
        ws.cell(row=row_num, column=2, value=visit.pharmacy_name)
        ws.cell(row=row_num, column=3, value=visit.pharmacy_address or "")
        ws.cell(row=row_num, column=4, value=visit.product_count)
        ws.cell(row=row_num, column=5, value=visit.mf_count)
        ws.cell(row=row_num, column=6, value=visit.visit_date.strftime('%d.%m.%Y') if visit.visit_date else "")
        ws.cell(row=row_num, column=7, value=visit.start_time.strftime('%H:%M') if visit.start_time else "")
        ws.cell(row=row_num, column=8, value=visit.end_time.strftime('%H:%M') if visit.end_time else "")
        ws.cell(row=row_num, column=9, value=visit.notes or "")
        ws.cell(row=row_num, column=10, value="Onaylandı" if visit.is_approved else "Bekliyor")

    # Sütun genişliklerini ayarla
    column_widths = [20, 30, 40, 12, 12, 15, 15, 15, 40, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # Excel'i hafızada oluştur
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Dosya adı oluştur - Türkçe ay isimleri
    turkish_months = {
        1: 'ocak', 2: 'subat', 3: 'mart', 4: 'nisan', 5: 'mayis', 6: 'haziran',
        7: 'temmuz', 8: 'agustos', 9: 'eylul', 10: 'ekim', 11: 'kasim', 12: 'aralik'
    }

    # Tarih formatı oluştur
    if start_date and end_date:
        start_month = turkish_months[start_date.month]
        start_day = start_date.day
        end_month = turkish_months[end_date.month]
        end_day = end_date.day
        year = end_date.year

        # Aynı gün mü kontrol et
        if start_date == end_date:
            date_str = f"{start_month}{start_day}_{year}"
        else:
            date_str = f"{start_month}{start_day}_{end_month}{end_day}_{year}"
    else:
        # Tarih yoksa bugünün tarihini kullan
        today = datetime.now()
        month = turkish_months[today.month]
        day = today.day
        year = today.year
        date_str = f"{month}{day}_{year}"

    filename = f"eczaneziyaretleri_{employee_name}_{date_str}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/pharmacies")
def get_pharmacy_visits(
    visit_date: Optional[date] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    employee_id: Optional[int] = None,
    pharmacy_id: Optional[int] = None,
    pharmacy_name: Optional[str] = None,  # Yeni: Eczane adı filtresi
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Eczane ziyaretlerini listele
    """
    query = db.query(PharmacyVisit)

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        query = query.filter(PharmacyVisit.employee_id == current_user.id)
    elif employee_id:
        query = query.filter(PharmacyVisit.employee_id == employee_id)

    # Pharmacy filtresi
    if pharmacy_id:
        query = query.filter(PharmacyVisit.pharmacy_id == pharmacy_id)

    # Eczane adı filtresi (case-insensitive, partial match)
    if pharmacy_name:
        search_term = f"%{pharmacy_name.lower().strip()}%"
        query = query.filter(PharmacyVisit.pharmacy_name.ilike(search_term))

    # Tarih filtresi - start_date ve end_date varsa aralık olarak filtrele
    if start_date and end_date:
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        query = query.filter(
            and_(
                PharmacyVisit.visit_date >= start_datetime,
                PharmacyVisit.visit_date <= end_datetime
            )
        )
    elif visit_date:
        # Sadece visit_date varsa tek tarih filtresi
        query = query.filter(PharmacyVisit.visit_date == visit_date)

    visits = query.order_by(PharmacyVisit.visit_date.desc()).offset(skip).limit(limit).all()

    # Employee adını ekle
    result = []
    for visit in visits:
        visit_dict = {
            "id": visit.id,
            "employee_id": visit.employee_id,
            "pharmacy_id": visit.pharmacy_id,
            "pharmacy_name": visit.pharmacy_name,
            "pharmacy_address": visit.pharmacy_address,
            "start_time": visit.start_time.isoformat() if visit.start_time else None,
            "end_time": visit.end_time.isoformat() if visit.end_time else None,
            "product_count": visit.product_count,
            "mf_count": visit.mf_count,
            "notes": visit.notes,
            "visit_date": visit.visit_date,
            "created_at": visit.created_at,
            "employee_name": visit.employee.full_name if visit.employee else None,
            "is_approved": visit.is_approved
        }
        result.append(visit_dict)

    return result



@router.get("/pharmacies/stats")
def get_pharmacy_visit_stats(
    period: Optional[str] = None,  # 'day', 'week', 'month', 'year'
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Eczane ziyaretleri istatistikleri
    """
    # Tarih aralığını belirle
    if not start_date or not end_date:
        # Period'a göre tarih aralığı belirle
        end_date_calc = datetime.now().date()
        start_date_calc = None

        if period == 'day':
            start_date_calc = end_date_calc
        elif period == 'week':
            start_date_calc = end_date_calc - timedelta(days=7)
        elif period == 'month':
            start_date_calc = end_date_calc - timedelta(days=30)
        elif period == 'year':
            start_date_calc = end_date_calc - timedelta(days=365)

        if not start_date:
            start_date = start_date_calc
        if not end_date:
            end_date = end_date_calc

    # Yetki kontrolü
    target_employee_id = None
    if current_user.role == EmployeeRole.EMPLOYEE:
        target_employee_id = current_user.id
    elif employee_id:
        target_employee_id = employee_id

    # Base query
    query = db.query(PharmacyVisit)

    if target_employee_id:
        query = query.filter(PharmacyVisit.employee_id == target_employee_id)
    if start_date:
        start_datetime = datetime.combine(start_date, datetime.min.time())
        query = query.filter(PharmacyVisit.visit_date >= start_datetime)
    if end_date:
        end_datetime = datetime.combine(end_date, datetime.max.time())
        query = query.filter(PharmacyVisit.visit_date <= end_datetime)

    # Toplam ziyaret sayısı
    total_visits = query.count()

    # Toplam MF sayısı
    mf_query = db.query(func.sum(PharmacyVisit.mf_count))
    if target_employee_id:
        mf_query = mf_query.filter(PharmacyVisit.employee_id == target_employee_id)
    if start_date:
        start_datetime = datetime.combine(start_date, datetime.min.time())
        mf_query = mf_query.filter(PharmacyVisit.visit_date >= start_datetime)
    if end_date:
        end_datetime = datetime.combine(end_date, datetime.max.time())
        mf_query = mf_query.filter(PharmacyVisit.visit_date <= end_datetime)
    total_mf = mf_query.scalar() or 0

    # Toplam satılan ürün sayısı
    products_query = db.query(func.sum(PharmacyVisit.product_count))
    if target_employee_id:
        products_query = products_query.filter(PharmacyVisit.employee_id == target_employee_id)
    if start_date:
        start_datetime = datetime.combine(start_date, datetime.min.time())
        products_query = products_query.filter(PharmacyVisit.visit_date >= start_datetime)
    if end_date:
        end_datetime = datetime.combine(end_date, datetime.max.time())
        products_query = products_query.filter(PharmacyVisit.visit_date <= end_datetime)
    total_products = products_query.scalar() or 0

    # Onaylanan / Onay bekleyen
    approved_count = query.filter(PharmacyVisit.is_approved == True).count()
    pending_count = query.filter(PharmacyVisit.is_approved == False).count()

    return {
        "total_visits": total_visits,
        "total_mf": int(total_mf),
        "total_products": int(total_products),
        "approved_count": approved_count,
        "pending_count": pending_count
    }

@router.post("/pharmacies", response_model=PharmacyVisitResponse)
def create_pharmacy_visit(
    visit: PharmacyVisitCreate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Yeni eczane ziyareti ekle
    """
    db_visit = PharmacyVisit(
        employee_id=current_user.id,
        **visit.dict()
    )

    db.add(db_visit)
    db.commit()
    db.refresh(db_visit)
    return db_visit


@router.get("/pharmacies/{visit_id}", response_model=PharmacyVisitResponse)
def get_pharmacy_visit(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Belirli bir eczane ziyaretini getir
    """
    visit = db.query(PharmacyVisit).filter(PharmacyVisit.id == visit_id).first()

    if not visit:
        raise HTTPException(status_code=404, detail="Visit not found")

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        if visit.employee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized")

    return visit


@router.delete("/pharmacies/{visit_id}")
def delete_pharmacy_visit(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Eczane ziyaretini sil
    """
    db_visit = db.query(PharmacyVisit).filter(PharmacyVisit.id == visit_id).first()

    if not db_visit:
        raise HTTPException(status_code=404, detail="Visit not found")

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        if db_visit.employee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized")

    db.delete(db_visit)
    db.commit()
    return {"message": "Visit deleted successfully"}


@router.put("/pharmacies/{visit_id}", response_model=PharmacyVisitResponse)
def update_pharmacy_visit(
    visit_id: int,
    visit_data: PharmacyVisitCreate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Eczane ziyaretini güncelle - Sadece aynı gün 23:59'a kadar düzenlenebilir
    """
    db_visit = db.query(PharmacyVisit).filter(PharmacyVisit.id == visit_id).first()

    if not db_visit:
        raise HTTPException(status_code=404, detail="Ziyaret bulunamadı")

    # Zaman kontrolü - Sadece aynı gün içinde (23:59'a kadar) düzenlenebilir
    now = datetime.now()
    visit_date_end = datetime.combine(db_visit.visit_date, datetime.max.time())  # Visit günü 23:59:59

    # EMPLOYEE ise: Sadece kendi ziyaretini ve sadece aynı gün içinde düzenleyebilir
    if current_user.role == EmployeeRole.EMPLOYEE:
        if db_visit.employee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Bu ziyaret size ait değil")

        if now > visit_date_end:
            raise HTTPException(
                status_code=403,
                detail="Ziyareti düzenlemek için son tarih geçti. Sadece ziyaret günü 23:59'a kadar düzenleyebilirsiniz."
            )

    # MANAGER/ADMIN ise: Herhangi bir ziyareti düzenleyebilir (zaman sınırı yok)
    elif current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    # Güncelle - Tüm güncellenebilir field'ları kaydet
    update_data = visit_data.dict(exclude_unset=True, exclude={'visit_date'})
    for field, value in update_data.items():
        setattr(db_visit, field, value)

    db.commit()
    db.refresh(db_visit)
    return db_visit


@router.post("/pharmacies/{visit_id}/toggle-approval")
def toggle_pharmacy_visit_approval(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Eczane ziyareti onayını toggle et (Manager/Admin only)
    """
    # Yetki kontrolü - Sadece Manager ve Admin onaylayabilir
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    # Ziyareti bul
    db_visit = db.query(PharmacyVisit).filter(PharmacyVisit.id == visit_id).first()
    if not db_visit:
        raise HTTPException(status_code=404, detail="Ziyaret bulunamadı")

    # Toggle approval
    db_visit.is_approved = not db_visit.is_approved
    db.commit()
    db.refresh(db_visit)

    return {
        "id": db_visit.id,
        "pharmacy_name": db_visit.pharmacy_name,
        "is_approved": db_visit.is_approved,
        "message": "Onay durumu güncellendi"
    }
