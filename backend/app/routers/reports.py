from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from typing import List, Optional
from datetime import date, datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import calendar

from ..database import get_db
from ..models.employee import Employee, EmployeeRole
from ..models.daily_report import DailyReport
from ..models.doctor_visit import DoctorVisit
from ..models.pharmacy_visit import PharmacyVisit
from ..models.sale import Sale
from ..models.weekly_program import WeeklyProgram
from ..schemas.report import DailyReportCreate, DailyReportUpdate, DailyReportResponse
from ..utils.dependencies import get_current_user

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/daily", response_model=List[DailyReportResponse])
def get_daily_reports(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Retrieve daily reports
    """
    query = db.query(DailyReport)

    # Filter by employee if not admin
    if current_user.role.value != "admin":
        query = query.filter(DailyReport.employee_id == current_user.id)

    reports = query.order_by(DailyReport.report_date.desc()).offset(skip).limit(limit).all()
    return reports


@router.get("/daily/{report_date}", response_model=DailyReportResponse)
def get_daily_report(
    report_date: date,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Get daily report by date
    """
    report = db.query(DailyReport).filter(
        DailyReport.employee_id == current_user.id,
        DailyReport.report_date == report_date
    ).first()

    if not report:
        raise HTTPException(status_code=404, detail="Daily report not found")

    return report


@router.post("/daily", response_model=DailyReportResponse)
def create_daily_report(
    report: DailyReportCreate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Create daily report with auto-calculated stats
    """
    # Check if report already exists for this date
    existing_report = db.query(DailyReport).filter(
        DailyReport.employee_id == current_user.id,
        DailyReport.report_date == report.report_date
    ).first()

    if existing_report:
        raise HTTPException(status_code=400, detail="Report already exists for this date")

    # Calculate stats
    start_datetime = datetime.combine(report.report_date, datetime.min.time())
    end_datetime = datetime.combine(report.report_date, datetime.max.time())

    # Count doctor and pharmacy visits
    doctor_visits = db.query(func.count(DoctorVisit.id)).filter(
        DoctorVisit.employee_id == current_user.id,
        DoctorVisit.visit_date >= start_datetime,
        DoctorVisit.visit_date <= end_datetime
    ).scalar() or 0

    pharmacy_visits = db.query(func.count(PharmacyVisit.id)).filter(
        PharmacyVisit.employee_id == current_user.id,
        PharmacyVisit.visit_date >= start_datetime,
        PharmacyVisit.visit_date <= end_datetime
    ).scalar() or 0

    total_visits = doctor_visits + pharmacy_visits

    total_sales = db.query(func.count(Sale.id)).filter(
        Sale.employee_id == current_user.id,
        Sale.sale_date >= start_datetime,
        Sale.sale_date <= end_datetime
    ).scalar()

    db_report = DailyReport(
        employee_id=current_user.id,
        total_visits=total_visits,
        total_sales=total_sales,
        **report.dict()
    )

    db.add(db_report)
    db.commit()
    db.refresh(db_report)
    return db_report


@router.put("/daily/{report_id}", response_model=DailyReportResponse)
def update_daily_report(
    report_id: int,
    report_update: DailyReportUpdate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Update daily report
    """
    db_report = db.query(DailyReport).filter(DailyReport.id == report_id).first()
    if not db_report:
        raise HTTPException(status_code=404, detail="Report not found")

    # Check permissions
    if current_user.role.value != "admin" and db_report.employee_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to update this report")

    update_data = report_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_report, field, value)

    db.commit()
    db.refresh(db_report)
    return db_report


@router.get("/export")
def export_report(
    period: str = Query(..., description="day, week, month, or year"),
    employee: Optional[str] = Query(None, description="Employee name or 'all'"),
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Export report data to Excel file
    """
    # Only managers can export reports
    if current_user.role != EmployeeRole.MANAGER:
        raise HTTPException(status_code=403, detail="Only managers can export reports")

    # Calculate date range based on period
    today = datetime.now().date()
    if period == 'day':
        start_date = today
        end_date = today
    elif period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif period == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    elif period == 'year':
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        raise HTTPException(status_code=400, detail="Invalid period")

    # For now, create mock data for demo
    # In production, this would query the actual database
    all_data = []

    # Mock doctor visits
    all_data.append({
        'date': '2025-11-24',
        'employee': 'Ahmet Kaya',
        'type': 'Doktor',
        'target': 'Dr. Mehmet Yılmaz',
        'specialty': 'Kardiyoloji',
        'hospital': 'Ankara Şehir Hastanesi',
        'start_time': '09:00',
        'end_time': '09:30',
        'status': 'Başarılı',
        'boxes_sold': 0,
        'boxes_gifted': 0,
        'feedback': 'Çok başarılı bir görüşme oldu.',
        'notes': 'Doktor ürünlerimizle ilgileniyor. Önümüzdeki ay tekrar ziyaret edilecek.'
    })

    all_data.append({
        'date': '2025-11-24',
        'employee': 'Ayşe Demir',
        'type': 'Doktor',
        'target': 'Dr. Zeynep Kara',
        'specialty': 'Nöroloji',
        'hospital': 'İstanbul Üniversitesi Hastanesi',
        'start_time': '14:00',
        'end_time': '14:45',
        'status': 'Başarılı',
        'boxes_sold': 0,
        'boxes_gifted': 0,
        'feedback': '',
        'notes': 'İlk görüşme. Bilgilendirme yapıldı.'
    })

    # Mock pharmacy visits
    all_data.append({
        'date': '2025-11-24',
        'employee': 'Ahmet Kaya',
        'type': 'Eczane',
        'target': 'Sağlık Eczanesi',
        'specialty': 'Eczacı Ali Veli',
        'hospital': 'Kızılay Cad. No:45 Ankara',
        'start_time': '10:00',
        'end_time': '10:30',
        'status': 'Anlaşma Sağlandı',
        'boxes_sold': 50,
        'boxes_gifted': 5,
        'feedback': 'Satış başarılı',
        'notes': 'Eczane ile anlaşma sağlandı, 50 kutu satış yapıldı.'
    })

    all_data.append({
        'date': '2025-11-23',
        'employee': 'Mehmet Öz',
        'type': 'Eczane',
        'target': 'Hayat Eczanesi',
        'specialty': 'Eczacı Fatma Şahin',
        'hospital': 'Atatürk Bulvarı No:123 İstanbul',
        'start_time': '15:00',
        'end_time': '15:20',
        'status': 'Anlaşma Sağlanamadı',
        'boxes_sold': 0,
        'boxes_gifted': 0,
        'feedback': 'Fiyat konusunda anlaşamadık',
        'notes': 'Eczacı başka tedarikçi ile çalışmak istiyor.'
    })

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Detaylı Rapor"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Tarih", "Çalışan", "Tür", "Hedef Adı", "Uzmanlık/Eczacı", "Hastane/Adres",
        "Başlangıç", "Bitiş", "Durum", "Satılan Kutu", "Hediye Kutu",
        "Manager Geri Dönüşü", "Notlar"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    for row_num, data in enumerate(all_data, 2):
        ws.cell(row=row_num, column=1, value=data['date'])
        ws.cell(row=row_num, column=2, value=data['employee'])
        ws.cell(row=row_num, column=3, value=data['type'])
        ws.cell(row=row_num, column=4, value=data['target'])
        ws.cell(row=row_num, column=5, value=data['specialty'])
        ws.cell(row=row_num, column=6, value=data['hospital'])
        ws.cell(row=row_num, column=7, value=data['start_time'])
        ws.cell(row=row_num, column=8, value=data['end_time'])
        ws.cell(row=row_num, column=9, value=data['status'])
        ws.cell(row=row_num, column=10, value=data['boxes_sold'])
        ws.cell(row=row_num, column=11, value=data['boxes_gifted'])
        ws.cell(row=row_num, column=12, value=data['feedback'])
        ws.cell(row=row_num, column=13, value=data['notes'])

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return as streaming response
    filename = f"rapor_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/weekly-plans")
def export_weekly_plans(
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Bu haftanın tüm çalışan planlarını export et
    """
    # Only managers/admins can export reports
    if current_user.role not in [EmployeeRole.MANAGER, EmployeeRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Only managers can export reports")

    # Bu haftanın başlangıç ve bitiş tarihlerini hesapla
    today = datetime.now().date()
    # Pazartesi'yi bul (hafta başlangıcı)
    start_of_week = today - timedelta(days=today.weekday())
    # Pazar'ı bul (hafta sonu)
    end_of_week = start_of_week + timedelta(days=6)

    # Bu haftanın programlarını çek
    programs = db.query(WeeklyProgram).filter(
        WeeklyProgram.week_start >= start_of_week,
        WeeklyProgram.week_start <= end_of_week
    ).all()

    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Haftalık Planlar"

    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=11)

    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"Haftalık Planlar ({start_of_week.strftime('%d.%m.%Y')} - {end_of_week.strftime('%d.%m.%Y')})"
    title_cell.fill = header_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    row_num = 3

    if not programs:
        ws['A3'] = "Bu hafta için plan bulunamadı"
    else:
        for program in programs:
            # Çalışan bilgisi
            employee = db.query(Employee).filter(Employee.id == program.employee_id).first()
            employee_name = employee.full_name if employee else "Bilinmeyen"

            ws.merge_cells(f'A{row_num}:E{row_num}')
            emp_cell = ws[f'A{row_num}']
            emp_cell.value = f"Çalışan: {employee_name}"
            emp_cell.fill = subheader_fill
            emp_cell.font = subheader_font
            emp_cell.alignment = Alignment(horizontal="left", vertical="center")
            row_num += 1

            # Günlük planlar
            days_json = program.days_json
            for day_data in days_json:
                day_name = day_data.get('day_name', '')
                day_date = day_data.get('date', '')
                visits = day_data.get('visits', [])

                if visits:
                    ws[f'A{row_num}'] = f"{day_name} ({day_date})"
                    ws[f'A{row_num}'].font = Font(bold=True)
                    row_num += 1

                    for visit in visits:
                        hospital = visit.get('hospital_name', '')
                        doctors = visit.get('doctors', [])

                        ws[f'B{row_num}'] = f"Hastane: {hospital}"
                        row_num += 1

                        for doctor in doctors:
                            ws[f'C{row_num}'] = f"• {doctor}"
                            row_num += 1

            row_num += 2  # Çalışanlar arası boşluk

    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"haftalik_planlar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/daily-reports")
def export_daily_reports(
    period: str = Query(..., description="day, week, month, or year"),
    employee: Optional[str] = Query(None, description="Employee name or 'all'"),
    visit_type: str = Query('all', description="all, doctor, or pharmacy"),
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Günlük raporları export et (doktor ve eczane ziyaretleri)
    """
    from ..models.leave_request import LeaveRequest, LeaveRequestStatus
    from ..models.leave_type import LeaveType

    # Only managers/admins can export reports
    if current_user.role not in [EmployeeRole.MANAGER, EmployeeRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Only managers can export reports")

    # Calculate date range
    today = datetime.now().date()
    if period == 'day':
        start_date = today
        end_date = today
    elif period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif period == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    elif period == 'year':
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        raise HTTPException(status_code=400, detail="Invalid period")

    # Get employees on leave for each date in the range
    def get_employee_leave_info(emp_id: int, check_date: date):
        """Check if employee is on leave for the given date"""
        leave = db.query(LeaveRequest, LeaveType).join(
            LeaveType, LeaveRequest.leave_type_id == LeaveType.id
        ).filter(
            LeaveRequest.employee_id == emp_id,
            LeaveRequest.status == LeaveRequestStatus.APPROVED,
            LeaveRequest.start_date <= check_date,
            LeaveRequest.end_date >= check_date
        ).first()

        if leave:
            leave_request, leave_type = leave
            return f"İzinli ({leave_type.name})"
        return None

    # Excel oluştur
    wb = Workbook()

    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Doktor ziyaretleri sheet (eğer visit_type all veya doctor ise)
    if visit_type in ['all', 'doctor']:
        ws_doctors = wb.active
        ws_doctors.title = "Hekim Ziyaretleri"

        # Query doctor visits
        doctor_query = db.query(DoctorVisit, Employee).join(
            Employee, DoctorVisit.employee_id == Employee.id
        ).filter(
            DoctorVisit.visit_date >= start_date,
            DoctorVisit.visit_date <= end_date
        )

        if employee and employee != 'all':
            doctor_query = doctor_query.filter(Employee.full_name == employee)

        doctor_visits = doctor_query.order_by(DoctorVisit.visit_date.desc()).all()

        # Headers
        headers = ["Tarih", "Çalışan", "Doktor Adı", "Hastane", "Branş", "Notlar"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_doctors.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data
        if doctor_visits:
            for row_num, (visit, emp) in enumerate(doctor_visits, 2):
                leave_info = get_employee_leave_info(emp.id, visit.visit_date.date())
                ws_doctors.cell(row=row_num, column=1, value=visit.visit_date.strftime('%d.%m.%Y'))
                ws_doctors.cell(row=row_num, column=2, value=emp.full_name)
                if leave_info:
                    ws_doctors.cell(row=row_num, column=3, value=leave_info)
                    ws_doctors.cell(row=row_num, column=4, value=leave_info)
                    ws_doctors.cell(row=row_num, column=5, value=leave_info)
                    ws_doctors.cell(row=row_num, column=6, value=leave_info)
                else:
                    ws_doctors.cell(row=row_num, column=3, value=visit.doctor_name)
                    ws_doctors.cell(row=row_num, column=4, value=visit.hospital_name)
                    ws_doctors.cell(row=row_num, column=5, value=visit.specialty or '')
                    ws_doctors.cell(row=row_num, column=6, value=visit.notes or '')
        else:
            # Veri yoksa mesaj ekle
            ws_doctors.merge_cells('A2:F2')
            no_data_cell = ws_doctors['A2']
            no_data_cell.value = "Bu dönem için hekim ziyareti kaydı bulunamadı"
            no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
            no_data_cell.font = Font(italic=True, color="999999")

        # Adjust column widths
        ws_doctors.column_dimensions['A'].width = 12
        ws_doctors.column_dimensions['B'].width = 20
        ws_doctors.column_dimensions['C'].width = 25
        ws_doctors.column_dimensions['D'].width = 30
        ws_doctors.column_dimensions['E'].width = 20
        ws_doctors.column_dimensions['F'].width = 40

    # Eczane ziyaretleri sheet (eğer visit_type all veya pharmacy ise)
    if visit_type in ['all', 'pharmacy']:
        if visit_type == 'pharmacy':
            ws_pharmacies = wb.active
            ws_pharmacies.title = "Eczane Ziyaretleri"
        else:
            ws_pharmacies = wb.create_sheet(title="Eczane Ziyaretleri")

        # Query pharmacy visits
        pharmacy_query = db.query(PharmacyVisit, Employee).join(
            Employee, PharmacyVisit.employee_id == Employee.id
        ).filter(
            PharmacyVisit.visit_date >= start_date,
            PharmacyVisit.visit_date <= end_date
        )

        if employee and employee != 'all':
            pharmacy_query = pharmacy_query.filter(Employee.full_name == employee)

        pharmacy_visits = pharmacy_query.order_by(PharmacyVisit.visit_date.desc()).all()

        # Headers
        headers = ["Tarih", "Çalışan", "Eczane Adı", "Satılan Ürün", "Verilen MF", "Notlar"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_pharmacies.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data
        if pharmacy_visits:
            for row_num, (visit, emp) in enumerate(pharmacy_visits, 2):
                leave_info = get_employee_leave_info(emp.id, visit.visit_date.date())
                ws_pharmacies.cell(row=row_num, column=1, value=visit.visit_date.strftime('%d.%m.%Y'))
                ws_pharmacies.cell(row=row_num, column=2, value=emp.full_name)
                if leave_info:
                    ws_pharmacies.cell(row=row_num, column=3, value=leave_info)
                    ws_pharmacies.cell(row=row_num, column=4, value=leave_info)
                    ws_pharmacies.cell(row=row_num, column=5, value=leave_info)
                    ws_pharmacies.cell(row=row_num, column=6, value=leave_info)
                else:
                    ws_pharmacies.cell(row=row_num, column=3, value=visit.pharmacy_name)
                    ws_pharmacies.cell(row=row_num, column=4, value=visit.product_count)
                    ws_pharmacies.cell(row=row_num, column=5, value=visit.mf_count)
                    ws_pharmacies.cell(row=row_num, column=6, value=visit.notes or '')
        else:
            # Veri yoksa mesaj ekle
            ws_pharmacies.merge_cells('A2:F2')
            no_data_cell = ws_pharmacies['A2']
            no_data_cell.value = "Bu dönem için eczane ziyareti kaydı bulunamadı"
            no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
            no_data_cell.font = Font(italic=True, color="999999")

        # Adjust column widths
        ws_pharmacies.column_dimensions['A'].width = 12
        ws_pharmacies.column_dimensions['B'].width = 20
        ws_pharmacies.column_dimensions['C'].width = 30
        ws_pharmacies.column_dimensions['D'].width = 15
        ws_pharmacies.column_dimensions['E'].width = 15
        ws_pharmacies.column_dimensions['F'].width = 40

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"gunluk_raporlar_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/growth-tracking")
def export_growth_tracking(
    employee: Optional[str] = Query(None, description="Employee name or 'all'"),
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Büyüme takibi raporu - Aylık bazda satış ve ziyaret performansı
    """
    # Only managers/admins can export reports
    if current_user.role not in [EmployeeRole.MANAGER, EmployeeRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Only managers can export reports")

    # Son 12 ayın verilerini al
    today = datetime.now().date()
    start_date = today - timedelta(days=365)

    # Employee filter
    employee_filter = None
    if employee and employee != 'all':
        emp = db.query(Employee).filter(Employee.full_name == employee).first()
        if emp:
            employee_filter = emp.id

    # Aylık verileri topla
    monthly_data = defaultdict(lambda: {
        'products': 0,
        'mf': 0,
        'doctor_visits': 0
    })

    # Eczane ziyaretlerinden ürün ve MF sayıları
    pharmacy_query = db.query(
        extract('year', PharmacyVisit.visit_date).label('year'),
        extract('month', PharmacyVisit.visit_date).label('month'),
        func.sum(PharmacyVisit.product_count).label('total_products'),
        func.sum(PharmacyVisit.mf_count).label('total_mf')
    ).filter(PharmacyVisit.visit_date >= start_date)

    if employee_filter:
        pharmacy_query = pharmacy_query.filter(PharmacyVisit.employee_id == employee_filter)

    pharmacy_stats = pharmacy_query.group_by(
        extract('year', PharmacyVisit.visit_date),
        extract('month', PharmacyVisit.visit_date)
    ).all()

    for stat in pharmacy_stats:
        key = f"{int(stat.year)}-{int(stat.month):02d}"
        monthly_data[key]['products'] = int(stat.total_products or 0)
        monthly_data[key]['mf'] = int(stat.total_mf or 0)

    # Doktor ziyaretleri
    doctor_query = db.query(
        extract('year', DoctorVisit.visit_date).label('year'),
        extract('month', DoctorVisit.visit_date).label('month'),
        func.count(DoctorVisit.id).label('total_visits')
    ).filter(DoctorVisit.visit_date >= start_date)

    if employee_filter:
        doctor_query = doctor_query.filter(DoctorVisit.employee_id == employee_filter)

    doctor_stats = doctor_query.group_by(
        extract('year', DoctorVisit.visit_date),
        extract('month', DoctorVisit.visit_date)
    ).all()

    for stat in doctor_stats:
        key = f"{int(stat.year)}-{int(stat.month):02d}"
        monthly_data[key]['doctor_visits'] = int(stat.total_visits or 0)

    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Büyüme Takibi"

    # Styling
    header_fill = PatternFill(start_color="9966CC", end_color="9966CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    growth_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    decline_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Büyüme Takibi Raporu - {employee if employee != 'all' else 'Tüm Çalışanlar'}"
    title_cell.fill = header_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Ay", "Satılan Ürün", "Verilen MF", "Hekim Ziyareti",
        "Ürün Değişim %", "MF Değişim %", "Ziyaret Değişim %", "3 Ay Ort. Değişim %"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ayları sırala ve yazdır
    sorted_months = sorted(monthly_data.keys())
    row_num = 4

    for i, month_key in enumerate(sorted_months):
        data = monthly_data[month_key]
        year, month = month_key.split('-')
        month_name = calendar.month_name[int(month)]

        ws.cell(row=row_num, column=1, value=f"{month_name} {year}")
        ws.cell(row=row_num, column=2, value=data['products'])
        ws.cell(row=row_num, column=3, value=data['mf'])
        ws.cell(row=row_num, column=4, value=data['doctor_visits'])

        # Önceki aya göre değişim
        if i > 0:
            prev_month = sorted_months[i - 1]
            prev_data = monthly_data[prev_month]

            # Ürün değişimi
            if prev_data['products'] > 0:
                product_change = ((data['products'] - prev_data['products']) / prev_data['products']) * 100
                cell = ws.cell(row=row_num, column=5, value=f"{product_change:.1f}%")
                cell.fill = growth_fill if product_change >= 0 else decline_fill

            # MF değişimi
            if prev_data['mf'] > 0:
                mf_change = ((data['mf'] - prev_data['mf']) / prev_data['mf']) * 100
                cell = ws.cell(row=row_num, column=6, value=f"{mf_change:.1f}%")
                cell.fill = growth_fill if mf_change >= 0 else decline_fill

            # Ziyaret değişimi
            if prev_data['doctor_visits'] > 0:
                visit_change = ((data['doctor_visits'] - prev_data['doctor_visits']) / prev_data['doctor_visits']) * 100
                cell = ws.cell(row=row_num, column=7, value=f"{visit_change:.1f}%")
                cell.fill = growth_fill if visit_change >= 0 else decline_fill

            # 3 aylık ortalama değişim
            if i >= 3:
                last_3_months = sorted_months[i-3:i]
                current_3_months = sorted_months[i-2:i+1]

                prev_avg = sum(monthly_data[m]['products'] + monthly_data[m]['mf'] + monthly_data[m]['doctor_visits']
                              for m in last_3_months) / 3
                current_avg = sum(monthly_data[m]['products'] + monthly_data[m]['mf'] + monthly_data[m]['doctor_visits']
                                 for m in current_3_months) / 3

                if prev_avg > 0:
                    avg_change = ((current_avg - prev_avg) / prev_avg) * 100
                    cell = ws.cell(row=row_num, column=8, value=f"{avg_change:.1f}%")
                    cell.fill = growth_fill if avg_change >= 0 else decline_fill

        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"buyume_takibi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
