from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import List, Optional
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

from ..database import get_db
from ..models.employee import Employee, EmployeeRole, Gender
from ..models.leave_type import LeaveType, GenderRestriction
from ..models.leave_balance import LeaveBalance
from ..models.leave_request import LeaveRequest, LeaveRequestStatus
from ..schemas.leave_request import (
    LeaveRequestCreate,
    LeaveRequestUpdate,
    LeaveRequestApprove,
    LeaveRequestResponse
)
from ..schemas.leave_balance import LeaveBalanceResponse
from ..utils.dependencies import get_current_user

router = APIRouter(prefix="/leave-requests", tags=["Leave Requests"])


def get_or_create_balance(db: Session, employee_id: int, leave_type_id: int, year: int) -> LeaveBalance:
    """İzin bakiyesi yoksa oluştur, varsa getir"""
    balance = db.query(LeaveBalance).filter(
        and_(
            LeaveBalance.employee_id == employee_id,
            LeaveBalance.leave_type_id == leave_type_id,
            LeaveBalance.year == year
        )
    ).first()

    if not balance:
        # Bakiye yoksa oluştur
        leave_type = db.query(LeaveType).filter(LeaveType.id == leave_type_id).first()
        employee = db.query(Employee).filter(Employee.id == employee_id).first()

        # Yıllık izin türü için hire_date'e göre hesapla
        if leave_type.name == "Yıllık İzin" and employee and employee.hire_date:
            from .annual_leave_rules import calculate_annual_leave_days
            total_days = calculate_annual_leave_days(employee.hire_date, year, db)
        else:
            # Diğer izin türleri için max_days kullan
            total_days = leave_type.max_days

        balance = LeaveBalance(
            employee_id=employee_id,
            leave_type_id=leave_type_id,
            year=year,
            total_days=total_days,
            used_days=0,
            remaining_days=total_days
        )
        db.add(balance)
        db.commit()
        db.refresh(balance)

    return balance


@router.get("/", response_model=List[LeaveRequestResponse])
def get_leave_requests(
    status_filter: Optional[LeaveRequestStatus] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    İzin taleplerini listele
    - Manager: Tüm talepleri görebilir
    - Employee: Sadece kendi taleplerini görebilir
    """
    query = db.query(LeaveRequest)

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        # Çalışan sadece kendi taleplerini görebilir
        query = query.filter(LeaveRequest.employee_id == current_user.id)
    elif employee_id:
        # Manager specific employee filtreleyebilir
        query = query.filter(LeaveRequest.employee_id == employee_id)

    # Status filtresi
    if status_filter:
        query = query.filter(LeaveRequest.status == status_filter)

    requests = query.order_by(LeaveRequest.created_at.desc()).all()

    # Response hazırla
    result = []
    for req in requests:
        employee = db.query(Employee).filter(Employee.id == req.employee_id).first()
        leave_type = db.query(LeaveType).filter(LeaveType.id == req.leave_type_id).first()
        approver = None
        if req.approved_by:
            approver = db.query(Employee).filter(Employee.id == req.approved_by).first()

        result.append(LeaveRequestResponse(
            id=req.id,
            employee_id=req.employee_id,
            employee_name=employee.full_name if employee else "Unknown",
            leave_type_id=req.leave_type_id,
            leave_type_name=leave_type.name if leave_type else "Unknown",
            start_date=req.start_date,
            end_date=req.end_date,
            return_to_work_date=req.return_to_work_date,
            total_days=req.total_days,
            status=req.status,
            message=req.message,
            rejection_reason=req.rejection_reason,
            approved_by=req.approved_by,
            approver_name=approver.full_name if approver else None,
            approved_at=req.approved_at,
            created_at=req.created_at,
            updated_at=req.updated_at
        ))

    return result


@router.get("/my-balances", response_model=List[LeaveBalanceResponse])
def get_my_leave_balances(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """Çalışanın izin bakiyelerini getir"""
    if not year:
        year = datetime.now().year

    # Tüm aktif izin türleri için bakiyeleri getir/oluştur
    active_leave_types = db.query(LeaveType).filter(LeaveType.is_active == True).all()

    balances = []
    for leave_type in active_leave_types:
        balance = get_or_create_balance(db, current_user.id, leave_type.id, year)
        balances.append(LeaveBalanceResponse(
            id=balance.id,
            employee_id=balance.employee_id,
            leave_type_id=balance.leave_type_id,
            leave_type_name=leave_type.name,
            year=balance.year,
            total_days=balance.total_days,
            used_days=balance.used_days,
            remaining_days=balance.remaining_days,
            created_at=balance.created_at,
            updated_at=balance.updated_at
        ))

    return balances


@router.post("/", response_model=LeaveRequestResponse, status_code=status.HTTP_201_CREATED)
def create_leave_request(
    request_data: LeaveRequestCreate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Yeni izin talebi oluştur
    Kontroller:
    - İzin türü aktif mi?
    - Cinsiyet kısıtına uyuyor mu?
    - Yeterli bakiye var mı?
    """
    # İzin türünü kontrol et
    leave_type = db.query(LeaveType).filter(LeaveType.id == request_data.leave_type_id).first()
    if not leave_type:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="İzin türü bulunamadı"
        )

    if not leave_type.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu izin türü aktif değil"
        )

    # Cinsiyet kısıtı kontrolü
    if leave_type.gender_restriction != GenderRestriction.NONE:
        if not current_user.gender:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cinsiyetiniz belirtilmemiş. Bu izin türü için cinsiyet bilgisi gereklidir."
            )

        if leave_type.gender_restriction == GenderRestriction.MALE_ONLY and current_user.gender != Gender.MALE:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Bu izin türü sadece erkek çalışanlar için geçerlidir"
            )

        if leave_type.gender_restriction == GenderRestriction.FEMALE_ONLY and current_user.gender != Gender.FEMALE:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Bu izin türü sadece kadın çalışanlar için geçerlidir"
            )

    # Bakiye kontrolü - Yıllık izin hariç diğerlerinde kontrol et
    year = request_data.start_date.year
    balance = get_or_create_balance(db, current_user.id, leave_type.id, year)

    # "Yıllık İzin" dışındaki izin türleri için bakiye kontrolü yap
    if leave_type.name != "Yıllık İzin" and balance.remaining_days < request_data.total_days:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Yetersiz bakiye. Kalan: {balance.remaining_days} gün, İstenilen: {request_data.total_days} gün"
        )

    # Talebi oluştur
    leave_request = LeaveRequest(
        employee_id=current_user.id,
        **request_data.model_dump()
    )
    db.add(leave_request)
    db.commit()
    db.refresh(leave_request)

    # Response hazırla
    return LeaveRequestResponse(
        id=leave_request.id,
        employee_id=leave_request.employee_id,
        employee_name=current_user.full_name,
        leave_type_id=leave_request.leave_type_id,
        leave_type_name=leave_type.name,
        start_date=leave_request.start_date,
        end_date=leave_request.end_date,
        return_to_work_date=leave_request.return_to_work_date,
        total_days=leave_request.total_days,
        status=leave_request.status,
        message=leave_request.message,
        rejection_reason=leave_request.rejection_reason,
        approved_by=leave_request.approved_by,
        approver_name=None,
        approved_at=leave_request.approved_at,
        created_at=leave_request.created_at,
        updated_at=leave_request.updated_at
    )


@router.put("/{request_id}", response_model=LeaveRequestResponse)
def update_leave_request(
    request_id: int,
    request_data: LeaveRequestUpdate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    İzin talebini güncelle
    - Sadece pending talepler güncellenebilir
    - Çalışan sadece kendi talebini, manager tümünü güncelleyebilir
    """
    # Talebi bul
    leave_request = db.query(LeaveRequest).filter(LeaveRequest.id == request_id).first()
    if not leave_request:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="İzin talebi bulunamadı"
        )

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        if leave_request.employee_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Bu talebi güncelleme yetkiniz yok"
            )

    # Sadece pending talepler güncellenebilir
    if leave_request.status != LeaveRequestStatus.PENDING:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sadece beklemedeki talepler güncellenebilir"
        )

    # Güncelle
    update_data = request_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(leave_request, field, value)

    db.commit()
    db.refresh(leave_request)

    # Response hazırla
    employee = db.query(Employee).filter(Employee.id == leave_request.employee_id).first()
    leave_type = db.query(LeaveType).filter(LeaveType.id == leave_request.leave_type_id).first()

    return LeaveRequestResponse(
        id=leave_request.id,
        employee_id=leave_request.employee_id,
        employee_name=employee.full_name,
        leave_type_id=leave_request.leave_type_id,
        leave_type_name=leave_type.name,
        start_date=leave_request.start_date,
        end_date=leave_request.end_date,
        return_to_work_date=leave_request.return_to_work_date,
        total_days=leave_request.total_days,
        status=leave_request.status,
        message=leave_request.message,
        rejection_reason=leave_request.rejection_reason,
        approved_by=leave_request.approved_by,
        approver_name=None,
        approved_at=leave_request.approved_at,
        created_at=leave_request.created_at,
        updated_at=leave_request.updated_at
    )


@router.post("/{request_id}/approve", response_model=LeaveRequestResponse)
def approve_or_reject_leave_request(
    request_id: int,
    approval_data: LeaveRequestApprove,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    İzin talebini onayla veya reddet (sadece Manager veya approve_leaves yetkisi olanlar)
    """
    from ..utils.dependencies import has_permission

    # Yetki kontrolü: Sadece MANAGER veya approve_leaves yetkisi olanlar
    if current_user.role != EmployeeRole.MANAGER and not has_permission(current_user, "approve_leaves"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="İzin onaylama yetkiniz yok. Bu işlem için MANAGER rolü veya 'approve_leaves' yetkisi gereklidir."
        )

    # Talebi bul
    leave_request = db.query(LeaveRequest).filter(LeaveRequest.id == request_id).first()
    if not leave_request:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="İzin talebi bulunamadı"
        )

    # Kendi talebini onaylayamaz
    if leave_request.employee_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Kendi izin talebinizi onaylayamazsınız"
        )

    # Sadece pending talepler onaylanabilir/reddedilebilir
    if leave_request.status != LeaveRequestStatus.PENDING:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sadece beklemedeki talepler işleme alınabilir"
        )

    if approval_data.approved:
        # ONAY
        leave_request.status = LeaveRequestStatus.APPROVED
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.utcnow()

        # Bakiyeyi güncelle
        year = leave_request.start_date.year
        balance = get_or_create_balance(db, leave_request.employee_id, leave_request.leave_type_id, year)
        balance.used_days += leave_request.total_days
        balance.remaining_days -= leave_request.total_days

    else:
        # RED
        if not approval_data.rejection_reason:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Red nedeni belirtilmelidir"
            )
        leave_request.status = LeaveRequestStatus.REJECTED
        leave_request.rejection_reason = approval_data.rejection_reason
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.utcnow()

    db.commit()
    db.refresh(leave_request)

    # Response hazırla
    employee = db.query(Employee).filter(Employee.id == leave_request.employee_id).first()
    leave_type = db.query(LeaveType).filter(LeaveType.id == leave_request.leave_type_id).first()

    return LeaveRequestResponse(
        id=leave_request.id,
        employee_id=leave_request.employee_id,
        employee_name=employee.full_name,
        leave_type_id=leave_request.leave_type_id,
        leave_type_name=leave_type.name,
        start_date=leave_request.start_date,
        end_date=leave_request.end_date,
        return_to_work_date=leave_request.return_to_work_date,
        total_days=leave_request.total_days,
        status=leave_request.status,
        message=leave_request.message,
        rejection_reason=leave_request.rejection_reason,
        approved_by=leave_request.approved_by,
        approver_name=current_user.full_name,
        approved_at=leave_request.approved_at,
        created_at=leave_request.created_at,
        updated_at=leave_request.updated_at
    )


@router.get("/active", response_model=List[LeaveRequestResponse])
def get_active_leave_requests(
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Aktif izinleri listele (başlamış ve henüz bitmemiş onaylanmış izinler)
    - Manager: Tüm aktif izinleri görebilir
    - Employee: Sadece kendi aktif izinlerini görebilir
    """
    today = date.today()

    query = db.query(LeaveRequest).filter(
        LeaveRequest.status == LeaveRequestStatus.APPROVED,
        LeaveRequest.start_date <= today,
        LeaveRequest.end_date >= today
    )

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        # Çalışan sadece kendi izinlerini görebilir
        query = query.filter(LeaveRequest.employee_id == current_user.id)

    requests = query.order_by(LeaveRequest.start_date.desc()).all()

    # Response hazırla
    result = []
    for req in requests:
        employee = db.query(Employee).filter(Employee.id == req.employee_id).first()
        leave_type = db.query(LeaveType).filter(LeaveType.id == req.leave_type_id).first()
        approver = None
        if req.approved_by:
            approver = db.query(Employee).filter(Employee.id == req.approved_by).first()

        result.append(LeaveRequestResponse(
            id=req.id,
            employee_id=req.employee_id,
            employee_name=employee.full_name if employee else "Unknown",
            leave_type_id=req.leave_type_id,
            leave_type_name=leave_type.name if leave_type else "Unknown",
            start_date=req.start_date,
            end_date=req.end_date,
            return_to_work_date=req.return_to_work_date,
            total_days=req.total_days,
            status=req.status,
            message=req.message,
            rejection_reason=req.rejection_reason,
            approved_by=req.approved_by,
            approver_name=approver.full_name if approver else None,
            approved_at=req.approved_at,
            created_at=req.created_at,
            updated_at=req.updated_at
        ))

    return result


@router.post("/{request_id}/cancel", response_model=LeaveRequestResponse)
def cancel_leave_request(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    İzin talebini iptal et
    - Manager herhangi bir talebi iptal edebilir (geçmiş tarihtekiler dahil)
    - Çalışan sadece kendi pending talebini iptal edebilir
    - Çalışanlar için: İzin başladıysa iptal edilemez
    - Manager için: Geçmiş tarihte bile iptal edilebilir, bakiye iade edilir
    """
    from ..utils.dependencies import has_permission

    # Talebi bul
    leave_request = db.query(LeaveRequest).filter(LeaveRequest.id == request_id).first()
    if not leave_request:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="İzin talebi bulunamadı"
        )

    # Yetki kontrolü
    is_manager_or_admin = current_user.role in [EmployeeRole.ADMIN, EmployeeRole.MANAGER] or has_permission(current_user, "approve_leaves")

    if not is_manager_or_admin:
        # Çalışan sadece kendi talebini iptal edebilir
        if leave_request.employee_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Bu talebi iptal etme yetkiniz yok"
            )
        # Ve sadece pending taleplerini iptal edebilir
        if leave_request.status != LeaveRequestStatus.PENDING:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Sadece beklemedeki talepleri iptal edebilirsiniz"
            )
        # İzin başlamış mı kontrol et
        if leave_request.start_date <= date.today():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="İzin başladığı için iptal edilemez"
            )

    # Manager için: APPROVED izinleri iptal edebilir, bakiyeyi geri iade eder
    # Eğer daha önce approved idiyse bakiyeyi geri al
    if leave_request.status == LeaveRequestStatus.APPROVED:
        year = leave_request.start_date.year
        balance = get_or_create_balance(db, leave_request.employee_id, leave_request.leave_type_id, year)
        balance.used_days -= leave_request.total_days
        balance.remaining_days += leave_request.total_days

    # İptal et
    leave_request.status = LeaveRequestStatus.CANCELLED
    db.commit()
    db.refresh(leave_request)

    # Response hazırla
    employee = db.query(Employee).filter(Employee.id == leave_request.employee_id).first()
    leave_type = db.query(LeaveType).filter(LeaveType.id == leave_request.leave_type_id).first()
    approver = None
    if leave_request.approved_by:
        approver = db.query(Employee).filter(Employee.id == leave_request.approved_by).first()

    return LeaveRequestResponse(
        id=leave_request.id,
        employee_id=leave_request.employee_id,
        employee_name=employee.full_name,
        leave_type_id=leave_request.leave_type_id,
        leave_type_name=leave_type.name,
        start_date=leave_request.start_date,
        end_date=leave_request.end_date,
        return_to_work_date=leave_request.return_to_work_date,
        total_days=leave_request.total_days,
        status=leave_request.status,
        message=leave_request.message,
        rejection_reason=leave_request.rejection_reason,
        approved_by=leave_request.approved_by,
        approver_name=approver.full_name if approver else None,
        approved_at=leave_request.approved_at,
        created_at=leave_request.created_at,
        updated_at=leave_request.updated_at
    )


@router.get("/my-leave-status")
def check_my_leave_status(
    check_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Çalışanın belirli bir tarihte izinli olup olmadığını kontrol et
    check_date verilmezse bugünü kontrol eder
    """
    target_date = check_date if check_date else date.today()

    leave = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id == current_user.id,
        LeaveRequest.status == LeaveRequestStatus.APPROVED,
        LeaveRequest.start_date <= target_date,
        LeaveRequest.end_date >= target_date
    ).first()

    if leave:
        leave_type = db.query(LeaveType).filter(LeaveType.id == leave.leave_type_id).first()
        return {
            "is_on_leave": True,
            "leave_type": leave_type.name if leave_type else "Bilinmeyen",
            "start_date": leave.start_date,
            "end_date": leave.end_date,
            "return_to_work_date": leave.return_to_work_date
        }

    return {"is_on_leave": False}


@router.get("/export")
def export_approved_leaves(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    employee_id: Optional[int] = None,
    employee_name: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Onaylanmış izinleri Excel'e export et
    Filtreler: employee_name, year, month
    """
    from ..utils.dependencies import has_permission
    from sqlalchemy import extract

    # Yetki kontrolü
    can_view_all = current_user.role in [EmployeeRole.ADMIN, EmployeeRole.MANAGER] or has_permission(current_user, "view_all_leaves")

    query = db.query(LeaveRequest).filter(
        LeaveRequest.status == LeaveRequestStatus.APPROVED
    )

    # Tarih filtreleri
    if start_date:
        query = query.filter(LeaveRequest.start_date >= start_date)
    if end_date:
        query = query.filter(LeaveRequest.end_date <= end_date)

    # Yıl ve ay filtresi
    if year:
        query = query.filter(extract('year', LeaveRequest.start_date) == year)
    if month:
        query = query.filter(extract('month', LeaveRequest.start_date) == month)

    # Çalışan filtresi (employee_name ile)
    if employee_name and can_view_all:
        employee = db.query(Employee).filter(Employee.full_name == employee_name).first()
        if employee:
            query = query.filter(LeaveRequest.employee_id == employee.id)
    elif employee_id and can_view_all:
        query = query.filter(LeaveRequest.employee_id == employee_id)
    elif not can_view_all:
        # Sadece kendi izinlerini görebilir
        query = query.filter(LeaveRequest.employee_id == current_user.id)

    leaves = query.order_by(LeaveRequest.start_date.desc()).all()

    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Onaylanmış İzinler"

    # Başlıklar
    headers = ["Çalışan", "İzin Türü", "Başlangıç Tarihi", "Bitiş Tarihi", "İşe Dönüş Tarihi", "Toplam Gün", "Onaylayan", "Onaylanma Tarihi"]
    ws.append(headers)

    # Başlık stilini ayarla
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Verileri ekle
    for leave in leaves:
        employee = db.query(Employee).filter(Employee.id == leave.employee_id).first()
        leave_type = db.query(LeaveType).filter(LeaveType.id == leave.leave_type_id).first()
        approver = None
        if leave.approved_by:
            approver = db.query(Employee).filter(Employee.id == leave.approved_by).first()

        ws.append([
            employee.full_name if employee else "Bilinmeyen",
            leave_type.name if leave_type else "Bilinmeyen",
            leave.start_date.strftime("%d.%m.%Y"),
            leave.end_date.strftime("%d.%m.%Y"),
            leave.return_to_work_date.strftime("%d.%m.%Y") if leave.return_to_work_date else "-",
            leave.total_days,
            approver.full_name if approver else "Bilinmeyen",
            leave.approved_at.strftime("%d.%m.%Y %H:%M") if leave.approved_at else "-"
        ])

    # Sütun genişliklerini ayarla
    column_widths = [20, 20, 15, 15, 18, 12, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Excel'i byte stream'e yaz
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Dosya adını oluştur
    from datetime import datetime as dt
    filename = f"onaylanmis_izinler_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.put("/{request_id}/edit-dates", response_model=LeaveRequestResponse)
def edit_approved_leave_dates(
    request_id: int,
    new_end_date: date = Query(..., description="Yeni bitiş tarihi"),
    new_return_date: date = Query(..., description="Yeni işe dönüş tarihi"),
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Manager'ın onaylanmış aktif izinlerin tarihlerini düzenlemesi
    Örnek: 40 günlük izin onaylanmış, çalışan 20. günde dönmek istiyor
    """
    # Yetki kontrolü - Sadece Manager yapabilir
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu işlem için yetkiniz yok"
        )

    # Talebi bul
    leave_request = db.query(LeaveRequest).filter(LeaveRequest.id == request_id).first()
    if not leave_request:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="İzin talebi bulunamadı"
        )

    # Sadece approved izinler düzenlenebilir
    if leave_request.status != LeaveRequestStatus.APPROVED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sadece onaylanmış izinler düzenlenebilir"
        )

    # Yeni bitiş tarihi başlangıçtan önce olamaz
    if new_end_date < leave_request.start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bitiş tarihi başlangıç tarihinden önce olamaz"
        )

    # İşe dönüş tarihi bitiş tarihinden önce olamaz
    if new_return_date < new_end_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="İşe dönüş tarihi bitiş tarihinden önce olamaz"
        )

    # Yeni gün sayısını hesapla
    old_total_days = leave_request.total_days
    new_total_days = (new_end_date - leave_request.start_date).days + 1

    # Bakiyeyi güncelle
    year = leave_request.start_date.year
    balance = get_or_create_balance(db, leave_request.employee_id, leave_request.leave_type_id, year)

    # Eski günleri geri ver
    balance.used_days -= old_total_days
    balance.remaining_days += old_total_days

    # Yeni günleri düş
    balance.used_days += new_total_days
    balance.remaining_days -= new_total_days

    # İzin talebini güncelle
    leave_request.end_date = new_end_date
    leave_request.return_to_work_date = new_return_date
    leave_request.total_days = new_total_days

    db.commit()
    db.refresh(leave_request)

    # Response hazırla
    employee = db.query(Employee).filter(Employee.id == leave_request.employee_id).first()
    leave_type = db.query(LeaveType).filter(LeaveType.id == leave_request.leave_type_id).first()
    approver = None
    if leave_request.approved_by:
        approver = db.query(Employee).filter(Employee.id == leave_request.approved_by).first()

    return LeaveRequestResponse(
        id=leave_request.id,
        employee_id=leave_request.employee_id,
        employee_name=employee.full_name,
        leave_type_id=leave_request.leave_type_id,
        leave_type_name=leave_type.name,
        start_date=leave_request.start_date,
        end_date=leave_request.end_date,
        return_to_work_date=leave_request.return_to_work_date,
        total_days=leave_request.total_days,
        status=leave_request.status,
        message=leave_request.message,
        rejection_reason=leave_request.rejection_reason,
        approved_by=leave_request.approved_by,
        approver_name=approver.full_name if approver else None,
        approved_at=leave_request.approved_at,
        created_at=leave_request.created_at,
        updated_at=leave_request.updated_at
    )


@router.get("/my-leave-status-today")
def check_my_leave_status_today(
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Çalışanın bugün izinli olup olmadığını kontrol et
    """
    target_date = date.today()

    leave = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id == current_user.id,
        LeaveRequest.status == LeaveRequestStatus.APPROVED,
        LeaveRequest.start_date <= target_date,
        LeaveRequest.end_date >= target_date
    ).first()

    if leave:
        leave_type = db.query(LeaveType).filter(LeaveType.id == leave.leave_type_id).first()
        return {
            "is_on_leave": True,
            "leave_type": leave_type.name if leave_type else "Bilinmeyen",
            "start_date": leave.start_date,
            "end_date": leave.end_date,
            "return_to_work_date": leave.return_to_work_date
        }

    return {"is_on_leave": False}


@router.get("/my-leave-status")
def check_my_leave_status(
    check_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Çalışanın belirli bir tarihte izinli olup olmadığını kontrol et
    check_date verilmezse bugünü kontrol eder
    """
    target_date = check_date if check_date else date.today()

    leave = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id == current_user.id,
        LeaveRequest.status == LeaveRequestStatus.APPROVED,
        LeaveRequest.start_date <= target_date,
        LeaveRequest.end_date >= target_date
    ).first()

    if leave:
        leave_type = db.query(LeaveType).filter(LeaveType.id == leave.leave_type_id).first()
        return {
            "is_on_leave": True,
            "leave_type": leave_type.name if leave_type else "Bilinmeyen",
            "start_date": leave.start_date,
            "end_date": leave.end_date,
            "return_to_work_date": leave.return_to_work_date
        }

    return {"is_on_leave": False}


@router.get("/export")
def export_approved_leaves(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    employee_id: Optional[int] = None,
    employee_name: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Onaylanmış izinleri Excel'e export et
    Filtreler: employee_name, year, month
    """
    from ..utils.dependencies import has_permission
    from sqlalchemy import extract

    # Yetki kontrolü
    can_view_all = current_user.role in [EmployeeRole.ADMIN, EmployeeRole.MANAGER] or has_permission(current_user, "view_all_leaves")

    query = db.query(LeaveRequest).filter(
        LeaveRequest.status == LeaveRequestStatus.APPROVED
    )

    # Tarih filtreleri
    if start_date:
        query = query.filter(LeaveRequest.start_date >= start_date)
    if end_date:
        query = query.filter(LeaveRequest.end_date <= end_date)

    # Yıl ve ay filtresi
    if year:
        query = query.filter(extract('year', LeaveRequest.start_date) == year)
    if month:
        query = query.filter(extract('month', LeaveRequest.start_date) == month)

    # Çalışan filtresi (employee_name ile)
    if employee_name and can_view_all:
        employee = db.query(Employee).filter(Employee.full_name == employee_name).first()
        if employee:
            query = query.filter(LeaveRequest.employee_id == employee.id)
    elif employee_id and can_view_all:
        query = query.filter(LeaveRequest.employee_id == employee_id)
    elif not can_view_all:
        # Sadece kendi izinlerini görebilir
        query = query.filter(LeaveRequest.employee_id == current_user.id)

    leaves = query.order_by(LeaveRequest.start_date.desc()).all()

    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Onaylanmış İzinler"

    # Başlıklar
    headers = ["Çalışan", "İzin Türü", "Başlangıç Tarihi", "Bitiş Tarihi", "İşe Dönüş Tarihi", "Toplam Gün", "Onaylayan", "Onaylanma Tarihi"]
    ws.append(headers)

    # Başlık stilini ayarla
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Verileri ekle
    for leave in leaves:
        employee = db.query(Employee).filter(Employee.id == leave.employee_id).first()
        leave_type = db.query(LeaveType).filter(LeaveType.id == leave.leave_type_id).first()
        approver = None
        if leave.approved_by:
            approver = db.query(Employee).filter(Employee.id == leave.approved_by).first()

        ws.append([
            employee.full_name if employee else "Bilinmeyen",
            leave_type.name if leave_type else "Bilinmeyen",
            leave.start_date.strftime("%d.%m.%Y"),
            leave.end_date.strftime("%d.%m.%Y"),
            leave.return_to_work_date.strftime("%d.%m.%Y") if leave.return_to_work_date else "-",
            leave.total_days,
            approver.full_name if approver else "Bilinmeyen",
            leave.approved_at.strftime("%d.%m.%Y %H:%M") if leave.approved_at else "-"
        ])

    # Sütun genişliklerini ayarla
    column_widths = [20, 20, 15, 15, 18, 12, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Excel'i byte stream'e yaz
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Dosya adını oluştur
    from datetime import datetime as dt
    filename = f"onaylanmis_izinler_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/employees-on-leave")
def get_employees_on_leave_for_date(
    check_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user)
):
    """
    Belirli bir tarihteki tüm izinli çalışanları getir (Manager/Admin için)
    """
    from ..utils.dependencies import has_permission

    # Yetki kontrolü
    if current_user.role not in [EmployeeRole.ADMIN, EmployeeRole.MANAGER] and not has_permission(current_user, "view_all_leaves"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu bilgilere erişim yetkiniz yok"
        )

    if not check_date:
        check_date = date.today()

    # O tarihteki izinli çalışanları bul
    leaves = db.query(LeaveRequest).filter(
        LeaveRequest.status == LeaveRequestStatus.APPROVED,
        LeaveRequest.start_date <= check_date,
        LeaveRequest.end_date >= check_date
    ).all()

    result = {}
    for leave in leaves:
        leave_type = db.query(LeaveType).filter(LeaveType.id == leave.leave_type_id).first()
        result[leave.employee_id] = {
            "is_on_leave": True,
            "leave_type": leave_type.name if leave_type else "Bilinmeyen",
            "start_date": str(leave.start_date),
            "end_date": str(leave.end_date)
        }

    return result
